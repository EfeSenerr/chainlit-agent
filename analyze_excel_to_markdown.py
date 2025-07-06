#!/usr/bin/env python3
"""
Excel Data Analyzer (First 18 Rows) and Markdown Generator

This script processes Excel files (.xlsx) in a specified directory and generates 
objective analyses using GPT-4o. It reads only the first 18 rows of each Excel file
to handle large datasets efficiently, analyzes the content, and creates detailed 
markdown reports with statistical insights.

Features:
- Processes .xlsx files in the specified directory
- Reads only the first 18 rows for manageable analysis
- Uses GPT-4o with Entra ID authentication for analysis
- Generates objective statistical analyses with numbers and trends
- Creates markdown files in data/excel_converted/ directory
- Handles complex Excel formats and structures
- Detailed progress reporting
- Robust error handling

Usage:
    python analyze_excel_to_markdown.py [directory_path] [target_file]
    
    If no directory is specified, it defaults to './data/product_info/data/'
    If target_file is specified, only that file will be processed
"""

import os
import sys
import re
from pathlib import Path
from typing import List, Dict, Any, Optional
from datetime import datetime
import pandas as pd
import openpyxl
from openpyxl import load_workbook

from azure.identity import DefaultAzureCredential, get_bearer_token_provider
from openai import AzureOpenAI
from dotenv import load_dotenv

# Load environment variables
load_dotenv()

class ExcelAnalyzer:
    """Analyzes Excel files (first 18 rows) and generates objective markdown reports"""
    
    def __init__(self):
        """Initialize the analyzer with Azure OpenAI client"""
        self.setup_openai_client()
        self.output_dir = Path("./data/excel_converted")
        self.output_dir.mkdir(parents=True, exist_ok=True)
        self.max_rows = 18  # Only read first 18 rows
        self.stats = {
            'files_processed': 0,
            'files_analyzed': 0,
            'analyses_generated': 0,
            'errors': 0
        }
    
    def setup_openai_client(self):
        """Setup Azure OpenAI client with Entra ID authentication"""
        try:
            api_base = os.getenv('AZURE_OPENAI_ENDPOINT')
            deployment_name = os.getenv('AZURE_OPENAI_CHAT_DEPLOYMENT', 'gpt-4o')
            api_version = "2024-02-01"  # Use known working API version
            
            if not api_base:
                raise ValueError("AZURE_OPENAI_ENDPOINT not configured in environment")
            
            # Create client using Entra ID authentication
            token_provider = get_bearer_token_provider(
                DefaultAzureCredential(), 
                "https://cognitiveservices.azure.com/.default"
            )
            
            self.client = AzureOpenAI(
                azure_endpoint=api_base,
                azure_ad_token_provider=token_provider,
                api_version=api_version
            )
            
            self.deployment_name = deployment_name
            print(f"✅ Azure OpenAI client initialized successfully")
            print(f"   🎯 Using deployment: {deployment_name}")
            print(f"   🔐 Authentication: Entra ID (DefaultAzureCredential)")
            print(f"   📅 API Version: {api_version}")
            
        except Exception as e:
            print(f"❌ Failed to initialize Azure OpenAI client: {e}")
            raise
    
    def inspect_excel_structure(self, file_path: Path) -> Dict[str, Any]:
        """
        Inspect Excel file structure to understand its layout
        
        Args:
            file_path: Path to the Excel file
            
        Returns:
            Dictionary containing file structure information
        """
        try:
            # Load workbook to get sheet information
            workbook = load_workbook(file_path, read_only=True, data_only=True)
            sheet_names = workbook.sheetnames
            
            # Get info about each sheet
            sheets_info = {}
            for sheet_name in sheet_names:
                try:
                    sheet = workbook[sheet_name]
                    # Get sheet dimensions (with safety checks)
                    max_row = sheet.max_row if sheet.max_row is not None else 0
                    max_col = sheet.max_column if sheet.max_column is not None else 0
                    
                    # Limit inspection to reasonable bounds
                    max_row = min(max_row, 50) if max_row > 0 else 0
                    max_col = min(max_col, 50) if max_col > 0 else 0
                    
                    sheets_info[sheet_name] = {
                        'max_row': max_row,
                        'max_col': max_col,
                        'dimensions': f"{max_row}x{max_col}"
                    }
                except Exception as sheet_error:
                    print(f"     ⚠️ Error inspecting sheet {sheet_name}: {sheet_error}")
                    sheets_info[sheet_name] = {
                        'max_row': 0,
                        'max_col': 0,
                        'dimensions': "unknown",
                        'error': str(sheet_error)
                    }
            
            workbook.close()
            
            return {
                'sheet_names': sheet_names,
                'sheets_info': sheets_info,
                'total_sheets': len(sheet_names)
            }
            
        except Exception as e:
            print(f"   ⚠️ Error inspecting Excel structure: {e}")
            # Fallback: try to get sheet names with pandas
            try:
                xl_file = pd.ExcelFile(file_path)
                sheet_names = xl_file.sheet_names
                xl_file.close()
                return {
                    'sheet_names': sheet_names,
                    'sheets_info': {name: {'dimensions': 'unknown'} for name in sheet_names},
                    'total_sheets': len(sheet_names)
                }
            except Exception as fallback_error:
                print(f"   ⚠️ Fallback also failed: {fallback_error}")
                return {'sheet_names': [], 'sheets_info': {}, 'total_sheets': 0}
    
    def read_excel_content(self, file_path: Path) -> Dict[str, Any]:
        """
        Read Excel file content (first 18 rows) and extract structured data
        
        Args:
            file_path: Path to the Excel file
            
        Returns:
            Dictionary containing Excel data and metadata
        """
        try:
            print(f"   📋 Inspecting Excel file structure...")
            
            # First, inspect the file structure
            structure_info = self.inspect_excel_structure(file_path)
            print(f"   📊 Found {structure_info['total_sheets']} sheet(s): {structure_info['sheet_names']}")
            
            # If no sheets found via openpyxl, try direct pandas approach
            if structure_info['total_sheets'] == 0:
                print(f"   🔄 Trying alternative Excel reading method...")
                try:
                    # Try to read with pandas ExcelFile
                    xl_file = pd.ExcelFile(file_path)
                    structure_info['sheet_names'] = xl_file.sheet_names
                    structure_info['total_sheets'] = len(xl_file.sheet_names)
                    xl_file.close()
                    print(f"   📊 Alternative method found {structure_info['total_sheets']} sheet(s): {structure_info['sheet_names']}")
                except Exception as e:
                    print(f"   ❌ Alternative method also failed: {e}")
                    return None
            
            # Read each sheet (first 18 rows only)
            sheets_data = {}
            
            for sheet_name in structure_info['sheet_names']:
                try:
                    print(f"   📄 Reading sheet: {sheet_name} (first {self.max_rows} rows)")
                    
                    # Try multiple approaches to read the sheet
                    df = None
                    
                    # Method 1: Read with pandas directly
                    try:
                        df = pd.read_excel(
                            file_path, 
                            sheet_name=sheet_name, 
                            nrows=self.max_rows,
                            header=None  # Don't assume first row is header
                        )
                        print(f"     ✅ Successfully read with pandas (Method 1)")
                    except Exception as e1:
                        print(f"     ⚠️ Method 1 failed: {e1}")
                        
                        # Method 2: Try with different encoding options
                        try:
                            df = pd.read_excel(
                                file_path, 
                                sheet_name=sheet_name, 
                                nrows=self.max_rows,
                                header=None,
                                engine='openpyxl'
                            )
                            print(f"     ✅ Successfully read with openpyxl engine (Method 2)")
                        except Exception as e2:
                            print(f"     ⚠️ Method 2 failed: {e2}")
                            
                            # Method 3: Try reading without row limit first, then limit
                            try:
                                df_full = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
                                df = df_full.head(self.max_rows)
                                print(f"     ✅ Successfully read with full-then-limit (Method 3)")
                            except Exception as e3:
                                print(f"     ❌ All methods failed: {e3}")
                                continue
                    
                    if df is None or df.empty:
                        print(f"     ⚠️ Empty or invalid data in sheet {sheet_name}")
                        continue
                    
                    # Try to detect if there's a proper header
                    potential_header_row = None
                    for i in range(min(5, len(df))):  # Check first 5 rows for headers
                        row_data = df.iloc[i].astype(str)
                        if any(not str(val).replace('.', '').replace(',', '').replace('-', '').isdigit() 
                              for val in row_data if pd.notna(val) and str(val).strip() and str(val) != 'nan'):
                            potential_header_row = i
                            break
                    
                    # If we found a potential header, re-read with proper header
                    if potential_header_row is not None and potential_header_row < len(df) - 1:
                        try:
                            df_with_header = pd.read_excel(
                                file_path, 
                                sheet_name=sheet_name, 
                                nrows=self.max_rows,
                                header=potential_header_row
                            )
                            if len(df_with_header.columns) > 1:  # Valid header found
                                df = df_with_header
                                print(f"     ✅ Header detected at row {potential_header_row}")
                        except:
                            pass  # Keep original df without header
                    
                    sheets_data[sheet_name] = {
                        'dataframe': df,
                        'shape': df.shape,
                        'columns': list(df.columns),
                        'header_row': potential_header_row
                    }
                    
                    print(f"     📊 Shape: {df.shape[0]} rows × {df.shape[1]} columns")
                    
                except Exception as e:
                    print(f"     ❌ Error reading sheet {sheet_name}: {e}")
                    continue
            
            if not sheets_data:
                print(f"   ❌ No sheets could be read from the Excel file")
                return None
            
            return {
                'sheets_data': sheets_data,
                'structure_info': structure_info,
                'file_name': file_path.name,
                'total_sheets_read': len(sheets_data)
            }
            
        except Exception as e:
            print(f"   ❌ Error reading Excel file: {e}")
            return None
    
    def analyze_excel_with_gpt4(self, excel_data: Dict[str, Any]) -> str:
        """
        Analyze Excel data using GPT-4o to generate objective analysis
        
        Args:
            excel_data: Dictionary containing Excel data and metadata
            
        Returns:
            Detailed objective analysis as markdown text
        """
        try:
            file_name = excel_data['file_name']
            sheets_data = excel_data['sheets_data']
            structure_info = excel_data['structure_info']
            
            # Prepare comprehensive data summary
            data_summary = f"""
File: {file_name}
Total Sheets: {excel_data['total_sheets_read']}
Sheet Names: {list(sheets_data.keys())}
Analysis Scope: First {self.max_rows} rows per sheet

SHEET-BY-SHEET ANALYSIS:
"""
            
            for sheet_name, sheet_info in sheets_data.items():
                df = sheet_info['dataframe']
                data_summary += f"""
--- SHEET: {sheet_name} ---
Shape: {df.shape[0]} rows × {df.shape[1]} columns
Header Row: {sheet_info.get('header_row', 'Not detected')}
Columns: {sheet_info['columns']}

Sample Data (first 10 rows):
{df.head(10).to_string()}

Data Types:
{df.dtypes.to_string()}

Non-null value counts:
{df.count().to_string()}

"""
            
            # Create comprehensive analysis prompt
            prompt = f"""Analyze this Excel dataset (first {self.max_rows} rows only) and provide a comprehensive, objective statistical analysis. The analysis should include:

## REQUIRED SECTIONS:

### 1. **File Overview**
- Excel file structure and organization
- Number of sheets and their purposes
- Data scope and limitations (first {self.max_rows} rows analysis)
- File organization and layout

### 2. **Data Structure Analysis**
- Sheet-by-sheet breakdown
- Column definitions and data types
- Header structure and organization
- Data completeness assessment

### 3. **Statistical Summary**
- Key numerical values and statistics from the data
- Important totals, counts, and percentages
- Notable patterns in the first {self.max_rows} rows
- Data ranges and distributions

### 4. **Content Analysis**
- What type of data this appears to be (demographic, statistical, administrative, etc.)
- Main categories or classifications present
- Geographic or temporal scope (if applicable)
- Purpose and context of the dataset

### 5. **Key Insights** (based on available data)
- Most significant findings from the analyzed portion
- Trends or patterns visible in the first {self.max_rows} rows
- Notable values or anomalies
- Implications for the full dataset

## ANALYSIS REQUIREMENTS:
- Be completely objective and factual
- Include ALL significant numbers and statistics found
- Clearly state the limitation (first {self.max_rows} rows only)
- Use proper statistical terminology
- Format as clear, professional markdown
- Include tables where appropriate
- Note what additional analysis might be possible with the full dataset

Dataset to analyze:
{data_summary}
"""

            # Call GPT-4o for analysis
            response = self.client.chat.completions.create(
                model=self.deployment_name,
                messages=[
                    {
                        "role": "system", 
                        "content": f"You are an expert data analyst specializing in Excel file analysis. Your task is to provide comprehensive, factual analyses of Excel datasets based on the first {self.max_rows} rows only. Always include specific numbers, percentages, and clear limitations of the partial analysis. Be thorough and professional while acknowledging the scope limitations."
                    },
                    {
                        "role": "user", 
                        "content": prompt
                    }
                ],
                max_tokens=2500,  # Increased for detailed analysis
                temperature=0.1   # Low temperature for objective analysis
            )
            
            analysis = response.choices[0].message.content.strip()
            
            # Add header with metadata
            header = f"""# Excel Data Analysis: {file_name.replace('.xlsx', '')}

**Generated:** {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}  
**Source File:** `{file_name}`  
**Analysis Scope:** First {self.max_rows} rows per sheet  
**Analysis Method:** GPT-4o Statistical Analysis  

---

## ⚠️ **Analysis Limitation Notice**
This analysis is based on **the first {self.max_rows} rows only** of each sheet in the Excel file. For comprehensive insights, a full dataset analysis would be required.

---

"""
            
            return header + analysis
                
        except Exception as e:
            error_msg = f"Failed to analyze Excel data: {str(e)}"
            print(f"   ⚠️ {error_msg}")
            return f"# Analysis Error\n\n**Error:** {error_msg}\n\n**File:** {excel_data['file_name']}"
    
    def process_excel_file(self, file_path: Path) -> bool:
        """
        Process a single Excel file to generate analysis
        
        Args:
            file_path: Path to the Excel file
            
        Returns:
            True if processing was successful, False otherwise
        """
        try:
            print(f"\n📄 Processing: {file_path.name}")
            
            # Read Excel data (first 18 rows)
            excel_data = self.read_excel_content(file_path)
            if excel_data is None:
                print(f"   ❌ Failed to read Excel data")
                self.stats['errors'] += 1
                return False
            
            print(f"   📊 Successfully read {excel_data['total_sheets_read']} sheet(s)")
            
            # Generate analysis with GPT-4o
            print(f"   🤖 Analyzing data with GPT-4o...")
            analysis = self.analyze_excel_with_gpt4(excel_data)
            
            # Create output file path
            output_filename = file_path.stem + ".md"
            output_path = self.output_dir / output_filename
            
            # Write analysis to markdown file
            with open(output_path, 'w', encoding='utf-8') as f:
                f.write(analysis)
            
            print(f"   ✅ Analysis saved to: {output_path}")
            
            self.stats['files_processed'] += 1
            self.stats['files_analyzed'] += 1
            self.stats['analyses_generated'] += 1
            
            return True
            
        except Exception as e:
            print(f"   ❌ Error processing {file_path.name}: {e}")
            self.stats['errors'] += 1
            return False
    
    def process_directory(self, directory_path: str, target_file: str = None) -> None:
        """
        Process Excel files in a directory
        
        Args:
            directory_path: Path to the directory to process
            target_file: Specific file to process (optional)
        """
        dir_path = Path(directory_path)
        
        if not dir_path.exists():
            print(f"❌ Directory does not exist: {directory_path}")
            return
        
        if target_file:
            # Process specific file
            target_path = dir_path / target_file
            if not target_path.exists():
                # Try adding .xlsx extension
                target_path = dir_path / f"{target_file}.xlsx"
            
            if target_path.exists() and target_path.suffix.lower() == '.xlsx':
                print(f"🎯 Processing specific file: {target_path.name}")
                self.process_excel_file(target_path)
            else:
                print(f"❌ Target file not found: {target_file}")
                return
        else:
            # Find all .xlsx files
            xlsx_files = list(dir_path.glob("**/*.xlsx"))
            
            if not xlsx_files:
                print(f"ℹ️ No .xlsx files found in {directory_path}")
                return
            
            print(f"🔍 Found {len(xlsx_files)} Excel files to process")
            print(f"📁 Processing directory: {dir_path.absolute()}")
            
            # Process each file
            for file_path in xlsx_files:
                self.process_excel_file(file_path)
        
        print(f"📁 Output directory: {self.output_dir.absolute()}")
        
        # Print summary
        self.print_summary()
    
    def print_summary(self) -> None:
        """Print processing summary statistics"""
        print(f"\n{'='*60}")
        print("📊 EXCEL ANALYSIS SUMMARY")
        print(f"{'='*60}")
        print(f"📄 Files processed: {self.stats['files_processed']}")
        print(f"📊 Files analyzed: {self.stats['files_analyzed']}")
        print(f"📝 Analyses generated: {self.stats['analyses_generated']}")
        print(f"📁 Output directory: {self.output_dir.absolute()}")
        print(f"⚠️ Analysis scope: First {self.max_rows} rows per sheet")
        if self.stats['errors'] > 0:
            print(f"❌ Errors encountered: {self.stats['errors']}")
        else:
            print("✅ No errors encountered")
        print(f"{'='*60}")

def main():
    """Main execution function"""
    print("🚀 Excel Data Analyzer (First 18 Rows) and Markdown Generator")
    print("=" * 60)
    
    # Parse command line arguments
    directory_path = "./data/product_info/data/"
    target_file = None
    
    if len(sys.argv) > 1:
        directory_path = sys.argv[1]
    if len(sys.argv) > 2:
        target_file = sys.argv[2]
    
    print(f"📁 Target directory: {Path(directory_path).absolute()}")
    if target_file:
        print(f"🎯 Target file: {target_file}")
    
    try:
        # Initialize analyzer
        analyzer = ExcelAnalyzer()
        
        # Process directory or specific file
        analyzer.process_directory(directory_path, target_file)
        
        print(f"\n🎉 ANALYSIS COMPLETED!")
        print("Next steps:")
        print("1. Review the generated analyses in data/excel_converted/")
        print("2. Verify the statistical insights and numbers")
        print("3. Consider running full dataset analysis if needed")
        print("4. Use the markdown files for your knowledge base indexing")
        
    except KeyboardInterrupt:
        print(f"\n⚠️ Operation cancelled by user")
    except Exception as e:
        print(f"\n❌ Operation failed: {e}")
        sys.exit(1)

if __name__ == "__main__":
    main()
